import openpyxl
import random
from openpyxl import Workbook, load_workbook

def wellness_question_data():
  path = "../dataset"
  wellness = path + "/wellness_dialog_script_dataset.xlsx"
  question = path + "/wellness_question.txt"

  f = open(question, 'w')

  wb = load_workbook(filename=wellness)
  ws = wb[wb.sheetnames[0]]
  for row in ws.iter_rows(min_row=2):
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()

def wellness_answer_data():
  path = "../dataset"
  wellness_dataset = path + "/wellness_dialog_script_dataset.xlsx"
  answer_dataset = path + "/wellness_answer.txt"

  f = open(answer_dataset, 'w')
  
  wb = load_workbook(filename=wellness_dataset)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows(min_row=2):
    if row[2].value == None: continue
    else: f.write(row[0].value + "    " + row[2].value + "\n")
  
  f.close()

def category_data():
  path = "../dataset"
  wellness_dataset = path + "/wellness_dialog_script_dataset.xlsx"
  category_dataset = path + "/wellness_category.txt"

  f = open(category_dataset, 'w')

  wb = load_workbook(filename=wellness_dataset)
  ws = wb[wb.sheetnames[0]]

  category_count = 0
  category_dict = []
  
  for row in ws.iter_rows(min_row=2):
    category = row[0].value
    if category not in category_dict:
      category_dict.append(category)
      f.write(category.strip() + "    " + str(category_count) + "\n")
      category_count += 1
      
  f.close()

def wellness_classification_data():
  path = "../data"
  category_dataset = path + "/wellness_category.txt"
  question_dataset = path + "/wellness_question.txt"
  classification_dataset = path + "/wellness_classification.txt"

  f_category = open(category_dataset, 'r')
  f_question = open(question_dataset, 'r')
  f_classification = open(classification_dataset, 'w')

  category_lines = f_category.readlines()
  category_dict = {}
  for _, data in enumerate(category_lines):
    data = data.split('    ')
    category_dict[data[0]] = data[1][:-1]

  question_lines = f_question.readlines()
  question_dict = {}
  for _, data in enumerate(question_lines):
    data = data.split('    ')
    f_classification.write(data[1][:-1] + "    " + category_dict[data[0]] + "\n")

  f_category.close()
  f_question.close()
  f_classification.close()
